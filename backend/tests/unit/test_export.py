import io
from datetime import date, datetime
from zoneinfo import ZoneInfo

import pytest
from openpyxl import load_workbook

from api.export.cells import (
    ExportCell,
    ExportContext,
    LocationInfo,
    is_date_only_due_date,
    location_title_by_kind,
    patient_cell,
    select_option_label,
    task_cell,
    to_zoned,
)
from api.export.labels import get_formats, get_labels
from api.export.render import (
    format_cell_text,
    neutralize_formula,
    render_csv,
    render_xlsx,
)
from api.export.service import _slugify_filename
from api.inputs import FieldType, PatientState, Sex
from database.models.patient import Patient
from database.models.property import PropertyDefinition, PropertyValue
from database.models.task import Task
from database.models.user import User

BERLIN = ZoneInfo("Europe/Berlin")


def make_context(locale: str = "de-DE") -> ExportContext:
    return ExportContext(
        labels=get_labels(locale),
        formats=get_formats(locale),
        tz=BERLIN,
        now=datetime(2026, 7, 5, 12, 0),
        exported_by="Erika Musterfrau",
    )


def make_patient(**overrides) -> Patient:
    values = {
        "id": "patient-1",
        "firstname": "Erika",
        "lastname": "Musterfrau",
        "birthdate": date(1960, 8, 1),
        "sex": Sex.FEMALE.value,
        "state": PatientState.ADMITTED.value,
        "clinic_id": "clinic-1",
        "position_id": "bed-1",
    }
    values.update(overrides)
    patient = Patient(**values)
    patient.tasks = overrides.get("tasks", [])
    return patient


def location_tree() -> dict[str, LocationInfo]:
    return {
        "clinic-1": LocationInfo("Innere Medizin", "CLINIC", None),
        "ward-1": LocationInfo("Station 3", "WARD", "clinic-1"),
        "room-1": LocationInfo("Zimmer 12", "ROOM", "ward-1"),
        "bed-1": LocationInfo("Bett A", "BED", "room-1"),
    }


def test_datetime_converted_to_berlin_and_german_format():
    ctx = make_context()
    cell = ExportCell(to_zoned(datetime(2026, 1, 5, 23, 30), ctx.tz), "datetime")
    assert format_cell_text(cell, ctx) == "06.01.2026 00:30"


def test_date_only_due_date_sentinel_keeps_utc_wall_date():
    assert is_date_only_due_date(datetime(2026, 3, 1, 23, 59, 59, 999000))
    assert not is_date_only_due_date(datetime(2026, 3, 1, 23, 59, 59))

    ctx = make_context()
    task = Task(id="t", title="x", done=False)
    task.due_date = datetime(2026, 3, 1, 23, 59, 59, 999000)
    task.assignees = []
    cell = task_cell(task, "dueDate", ctx)
    assert cell.kind == "date"
    assert format_cell_text(cell, ctx) == "01.03.2026"


def test_regular_due_date_converted_to_timezone():
    ctx = make_context()
    task = Task(id="t", title="x", done=False)
    task.due_date = datetime(2026, 7, 1, 10, 0)
    task.assignees = []
    cell = task_cell(task, "dueDate", ctx)
    assert cell.kind == "datetime"
    assert format_cell_text(cell, ctx) == "01.07.2026 12:00"


def test_task_assignee_names_and_team_fallback():
    ctx = make_context()
    ctx.locations = {"team-1": LocationInfo("Pflege Team", "TEAM", None)}
    task = Task(id="t", title="x", done=False)
    task.assignees = [
        User(id="u1", username="anna", firstname="Anna", lastname="Muster"),
        User(id="u2", username="jonas"),
    ]
    assert task_cell(task, "assignee", ctx).value == "Anna Muster, jonas"

    task.assignees = []
    task.assignee_team_id = "team-1"
    assert task_cell(task, "assignee", ctx).value == "Pflege Team"


def test_task_without_patient_uses_localized_label():
    ctx = make_context()
    task = Task(id="t", title="x", done=False)
    task.assignees = []
    task.patient = None
    assert task_cell(task, "patient", ctx).value == "Kein Patient"


def test_patient_cells_localized():
    ctx = make_context()
    ctx.locations = location_tree()
    done_task = Task(id="t1", title="a", done=True)
    open_task = Task(id="t2", title="b", done=False)
    patient = make_patient(tasks=[done_task, open_task])

    assert patient_cell(patient, "name", ctx).value == "Erika Musterfrau"
    assert patient_cell(patient, "state", ctx).value == "Aufgenommen"
    assert patient_cell(patient, "sex", ctx).value == "Weiblich"
    assert patient_cell(patient, "clinic", ctx).value == "Innere Medizin"
    assert patient_cell(patient, "position", ctx).value == "Bett A"
    assert patient_cell(patient, "location-WARD", ctx).value == "Station 3"
    assert patient_cell(patient, "location-ROOM", ctx).value == "Zimmer 12"
    assert patient_cell(patient, "location-BED", ctx).value == "Bett A"
    assert patient_cell(patient, "tasks", ctx).value == "1/2"
    assert (
        patient_cell(patient, "birthdate", ctx).value
        == "01.08.1960 (65 Jahre)"
    )


def test_discharged_patient_task_progress_is_zero():
    ctx = make_context()
    patient = make_patient(
        state=PatientState.DISCHARGED.value,
        tasks=[Task(id="t1", title="a", done=True)],
    )
    assert patient_cell(patient, "tasks", ctx).value == "0/0"


def _select_definition(definition_id: str, field_type: FieldType) -> PropertyDefinition:
    return PropertyDefinition(
        id=definition_id,
        name="Diet",
        field_type=field_type.value,
        options="Vegetarisch,Vegan,Normal",
    )


def test_select_option_label_resolves_option_keys():
    definition = _select_definition("def-1", FieldType.FIELD_TYPE_SELECT)
    assert select_option_label("def-1-opt-0", definition) == "Vegetarisch"
    assert select_option_label("def-1-opt-2", definition) == "Normal"
    assert select_option_label("def-1-opt-9", definition) == "def-1-opt-9"
    assert select_option_label("free text", definition) == "free text"


def test_select_property_cell_shows_option_label():
    ctx = make_context()
    definition = _select_definition("def-1", FieldType.FIELD_TYPE_SELECT)
    ctx.property_definitions = {"def-1": definition}
    ctx.properties_by_entity = {
        "patient-1": {
            "def-1": PropertyValue(
                id="pv-1",
                definition_id="def-1",
                patient_id="patient-1",
                select_value="def-1-opt-1",
            ),
        },
    }
    patient = make_patient()
    assert patient_cell(patient, "property_def-1", ctx).value == "Vegan"


def test_multi_select_property_cell_joins_option_labels_inline():
    ctx = make_context()
    definition = _select_definition("def-1", FieldType.FIELD_TYPE_MULTI_SELECT)
    ctx.property_definitions = {"def-1": definition}
    ctx.properties_by_entity = {
        "task-1": {
            "def-1": PropertyValue(
                id="pv-1",
                definition_id="def-1",
                task_id="task-1",
                multi_select_values="def-1-opt-0,def-1-opt-2",
            ),
        },
    }
    task = Task(id="task-1", title="x", done=False)
    task.assignees = []
    cell = task_cell(task, "property_def-1", ctx)
    assert cell.value == "Vegetarisch, Normal"


def test_location_title_by_kind_missing_levels():
    locations = {
        "ward-1": LocationInfo("Station 3", "WARD", None),
        "room-1": LocationInfo("Zimmer 12", "ROOM", "ward-1"),
    }
    assert (
        location_title_by_kind("room-1", ("WARD",), locations) == "Station 3"
    )
    assert location_title_by_kind("room-1", ("BED",), locations) is None
    assert location_title_by_kind(None, ("BED",), locations) is None


def test_number_formatting_uses_german_decimal_separator():
    ctx = make_context()
    assert format_cell_text(ExportCell(3.5, "number"), ctx) == "3,5"
    assert format_cell_text(ExportCell(42.0, "number"), ctx) == "42"
    english = make_context("en-US")
    assert format_cell_text(ExportCell(3.5, "number"), english) == "3.5"


def test_bool_formatting_is_localized():
    ctx = make_context()
    assert format_cell_text(ExportCell(True, "bool"), ctx) == "Ja"
    assert format_cell_text(ExportCell(False, "bool"), ctx) == "Nein"


def test_formula_injection_is_neutralized():
    assert neutralize_formula("=SUM(A1:A2)") == "'=SUM(A1:A2)"
    assert neutralize_formula("+49 123") == "'+49 123"
    assert neutralize_formula("@foo") == "'@foo"
    assert neutralize_formula("normal") == "normal"


def test_render_csv_uses_semicolon_bom_and_crlf():
    ctx = make_context()
    rows = [
        [ExportCell("Erika; Musterfrau"), ExportCell(True, "bool")],
        [ExportCell("=cmd"), ExportCell(None)],
    ]
    payload = render_csv(["Name", "Erledigt"], rows, ctx)
    assert payload.startswith(b"\xef\xbb\xbf")
    text = payload.decode("utf-8-sig")
    lines = text.split("\r\n")
    assert lines[0] == "Name;Erledigt"
    assert lines[1] == '"Erika; Musterfrau";Ja'
    assert lines[2] == "'=cmd;"


def test_render_xlsx_printable_layout():
    ctx = make_context()
    rows = [
        [
            ExportCell("Erika Musterfrau"),
            ExportCell(datetime(2026, 7, 1, 12, 0), "datetime"),
        ],
        [ExportCell("Max Mustermann"), ExportCell(None)],
    ]
    payload = render_xlsx(["Name", "Fällig"], rows, ctx, "Station 3")
    workbook = load_workbook(io.BytesIO(payload))
    sheet = workbook.active

    assert sheet.title == "Station 3"
    assert sheet.cell(row=1, column=1).value == "Station 3"
    subtitle = sheet.cell(row=2, column=1).value
    assert "von Erika Musterfrau" in subtitle
    assert sheet.cell(row=4, column=1).value == "Name"
    assert sheet.cell(row=5, column=1).value == "Erika Musterfrau"
    assert sheet.cell(row=5, column=2).value == datetime(2026, 7, 1, 12, 0)
    assert sheet.cell(row=5, column=2).number_format == "DD.MM.YYYY HH:MM"
    assert sheet.freeze_panes == "A5"
    assert sheet.print_title_rows == "$4:$4"
    assert sheet.page_setup.orientation == "landscape"
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.row_dimensions[4].height == 26
    assert sheet.row_dimensions[5].height >= 22


def test_slugify_filename_handles_umlauts():
    assert _slugify_filename("Meine Aufgaben – Station 3ä") == (
        "meine-aufgaben-station-3ae"
    )
    assert _slugify_filename("///") == "export"


@pytest.mark.asyncio
async def test_run_table_export_patients_end_to_end(
    db_session,
    sample_patient,
    sample_user_with_location_access,
):
    from api.context import Context
    from api.export.schemas import TableExportRequest
    from api.export.service import run_table_export

    context = Context(db=db_session, user=sample_user_with_location_access)
    request = TableExportRequest.model_validate({
        "format": "csv",
        "columns": [
            {"key": "name", "label": "Name"},
            {"key": "state", "label": "Status"},
            {"key": "clinic", "label": "Klinik"},
        ],
        "states": ["ADMITTED"],
        "locale": "de-DE",
        "timezone": "Europe/Berlin",
    })

    result = await run_table_export(context, "patients", request)

    assert result.filename.endswith(".csv")
    text = result.content.decode("utf-8-sig")
    lines = text.strip().split("\r\n")
    assert lines[0] == "Name;Status;Klinik"
    assert lines[1] == "John Doe;Aufgenommen;Test Clinic"


@pytest.mark.asyncio
async def test_run_table_export_tasks_xlsx_end_to_end(
    db_session,
    sample_task,
    sample_user_with_location_access,
):
    from api.context import Context
    from api.export.schemas import TableExportRequest
    from api.export.service import run_table_export

    context = Context(db=db_session, user=sample_user_with_location_access)
    request = TableExportRequest.model_validate({
        "format": "xlsx",
        "columns": [
            {"key": "title", "label": "Titel"},
            {"key": "patient", "label": "Patient"},
            {"key": "done", "label": "Erledigt"},
        ],
        "title": "Meine Aufgaben",
    })

    result = await run_table_export(context, "tasks", request)

    workbook = load_workbook(io.BytesIO(result.content))
    sheet = workbook.active
    assert sheet.cell(row=1, column=1).value == "Meine Aufgaben"
    assert "Test User" in sheet.cell(row=2, column=1).value
    assert sheet.cell(row=4, column=1).value == "Titel"
    assert sheet.cell(row=5, column=1).value == "Test Task"
    assert sheet.cell(row=5, column=2).value == "John Doe"
    assert sheet.cell(row=5, column=3).value == "Nein"
