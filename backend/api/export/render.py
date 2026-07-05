import codecs
import csv
import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

from api.export.cells import ExportCell, ExportContext

# Values starting with these characters would be interpreted as formulas by
# spreadsheet applications; they get an apostrophe prefix so patient data can
# never execute as a formula (CSV/formula injection).
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

_HEADER_FILL = PatternFill(
    start_color="FF334155", end_color="FF334155", fill_type="solid",
)
_ZEBRA_FILL = PatternFill(
    start_color="FFF1F5F9", end_color="FFF1F5F9", fill_type="solid",
)
_THIN_BORDER = Border(
    left=Side(style="thin", color="FFCBD5E1"),
    right=Side(style="thin", color="FFCBD5E1"),
    top=Side(style="thin", color="FFCBD5E1"),
    bottom=Side(style="thin", color="FFCBD5E1"),
)

_MIN_COLUMN_WIDTH = 10
_MAX_COLUMN_WIDTH = 45


def neutralize_formula(text: str) -> str:
    if text.startswith(_FORMULA_PREFIXES):
        return f"'{text}"
    return text


def format_number_text(value: float | int, ctx: ExportContext) -> str:
    if isinstance(value, bool):
        return format_cell_text(ExportCell(value, kind="bool"), ctx)
    if isinstance(value, int) or float(value).is_integer():
        return str(int(value))
    return str(value).replace(".", ctx.formats["decimal_separator"])


def format_cell_text(cell: ExportCell, ctx: ExportContext) -> str:
    value = cell.value
    if value is None:
        return ""
    if cell.kind == "bool":
        return ctx.labels["yes"] if value else ctx.labels["no"]
    if cell.kind == "datetime" and isinstance(value, datetime):
        return value.strftime(ctx.formats["datetime"])
    if cell.kind == "date":
        if isinstance(value, datetime):
            value = value.date()
        if isinstance(value, date):
            return value.strftime(ctx.formats["date"])
    if cell.kind == "number" and isinstance(value, (int, float)):
        return format_number_text(value, ctx)
    return neutralize_formula(str(value))


def render_csv(
    headers: list[str],
    rows: list[list[ExportCell]],
    ctx: ExportContext,
) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", lineterminator="\r\n")
    writer.writerow([neutralize_formula(header) for header in headers])
    for row in rows:
        writer.writerow([format_cell_text(cell, ctx) for cell in row])
    return codecs.BOM_UTF8 + buffer.getvalue().encode("utf-8")


def _sheet_title(title: str) -> str:
    cleaned = "".join(c for c in title if c not in "[]:*?/\\")
    return cleaned.strip()[:31] or "Export"


def _write_cell(worksheet, row: int, column: int, cell: ExportCell, ctx: ExportContext):
    target = worksheet.cell(row=row, column=column)
    value = cell.value
    if value is None:
        target.value = None
    elif cell.kind == "datetime" and isinstance(value, datetime):
        target.value = value
        target.number_format = ctx.formats["xlsx_datetime"]
    elif cell.kind == "date" and isinstance(value, (date, datetime)):
        target.value = value
        target.number_format = ctx.formats["xlsx_date"]
    elif cell.kind == "number" and isinstance(value, (int, float)):
        target.value = value
    elif cell.kind == "bool":
        target.value = ctx.labels["yes"] if value else ctx.labels["no"]
    else:
        target.value = neutralize_formula(str(value))
    return target


def _estimate_width(values: list[str]) -> float:
    longest = max((len(v) for v in values), default=0)
    return min(max(longest + 3, _MIN_COLUMN_WIDTH), _MAX_COLUMN_WIDTH)


def render_xlsx(
    headers: list[str],
    rows: list[list[ExportCell]],
    ctx: ExportContext,
    title: str,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _sheet_title(title)

    column_count = max(len(headers), 1)
    generated_at = ctx.now.strftime(ctx.formats["datetime"])
    subtitle = (
        f"{ctx.labels['generated_at']} {generated_at} ({ctx.tz.key}) — "
        f"{len(rows)} {ctx.labels['entries']}"
    )

    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=14, bold=True)
    worksheet.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=column_count,
    )
    subtitle_cell = worksheet.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = Font(size=9, color="FF64748B")
    worksheet.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=column_count,
    )

    header_row = 4
    for index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=index, value=header)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(
            vertical="center", wrap_text=True,
        )

    for row_offset, row in enumerate(rows):
        row_index = header_row + 1 + row_offset
        for column_index, cell in enumerate(row, start=1):
            written = _write_cell(worksheet, row_index, column_index, cell, ctx)
            written.border = _THIN_BORDER
            written.alignment = Alignment(vertical="top", wrap_text=True)
            if row_offset % 2 == 1:
                written.fill = _ZEBRA_FILL

    for column_index in range(1, column_count + 1):
        texts = [headers[column_index - 1]]
        for row in rows:
            if column_index <= len(row):
                texts.append(format_cell_text(row[column_index - 1], ctx))
        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = _estimate_width(texts)

    last_row = header_row + len(rows)
    worksheet.freeze_panes = f"A{header_row + 1}"
    if rows:
        worksheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(column_count)}{last_row}"
        )

    worksheet.print_title_rows = f"{header_row}:{header_row}"
    worksheet.print_area = (
        f"A1:{get_column_letter(column_count)}{max(last_row, header_row)}"
    )
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    worksheet.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.3, footer=0.3,
    )
    worksheet.oddFooter.left.text = f"{title} — {subtitle}"
    worksheet.oddFooter.left.size = 8
    worksheet.oddFooter.right.text = (
        f"{ctx.labels['page']} &P {ctx.labels['page_of']} &N"
    )
    worksheet.oddFooter.right.size = 8

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
